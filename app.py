"""
Budget Model Importer — Streamlit Web App
==========================================
Run:  streamlit run app.py
Requires: pip install streamlit openpyxl
"""

import io, json, os, re, statistics, tempfile, calendar, copy
from collections import defaultdict
from datetime import date, datetime

import openpyxl
from openpyxl.utils import get_column_letter
import streamlit as st

# ── Page config ───────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Budget Model Importer",
    page_icon="🏢",
    layout="centered",
    initial_sidebar_state="collapsed",
)

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap');
html, body, [class*="css"] { font-family: 'Inter', sans-serif; }
#MainMenu, footer, header { visibility: hidden; }
.block-container { padding-top: 2rem; padding-bottom: 3rem; max-width: 780px; }
.wordmark { display:flex; align-items:baseline; gap:10px; margin-bottom:2rem;
  padding-bottom:1.25rem; border-bottom:1px solid #E5E7EB; }
.wordmark-name { font-size:17px; font-weight:600; color:#111827; letter-spacing:-0.02em; }
.wordmark-tag  { font-size:12px; color:#9CA3AF; letter-spacing:0.05em; text-transform:uppercase; }
.step-label { font-size:11px; font-weight:600; letter-spacing:0.08em; text-transform:uppercase;
  color:#6B7280; margin-bottom:6px; margin-top:1.5rem; }
.info-card { background:#F9FAFB; border:1px solid #E5E7EB; border-radius:8px;
  padding:14px 16px; margin-bottom:1rem; font-size:13px; color:#374151; line-height:1.6; }
.val-row { display:flex; justify-content:space-between; align-items:center;
  padding:8px 0; border-bottom:1px solid #F3F4F6; font-size:13px; }
.val-row:last-child { border-bottom:none; }
.val-label { color:#6B7280; }
.val-value { font-weight:500; color:#111827; font-variant-numeric:tabular-nums; }
.success-banner { background:#ECFDF5; border:1px solid #A7F3D0; border-radius:8px;
  padding:12px 16px; font-size:14px; color:#065F46; font-weight:500; margin-bottom:1rem; }
.warn-banner { background:#FFFBEB; border:1px solid #FCD34D; border-radius:8px;
  padding:12px 16px; font-size:13px; color:#92400E; margin-bottom:0.5rem; }
div[data-testid="stDownloadButton"] > button {
  width:100%; background:#111827 !important; color:white !important; border:none !important;
  border-radius:7px !important; padding:10px 20px !important; font-weight:500 !important;
  font-size:14px !important; }
div[data-testid="stDownloadButton"] > button:hover { background:#374151 !important; }
div[data-testid="stButton"] > button[kind="primary"] {
  width:100%; background:#111827 !important; color:white !important; border:none !important;
  border-radius:7px !important; padding:10px 20px !important; font-weight:500 !important; }
</style>
""", unsafe_allow_html=True)

# ── Load pipeline modules ─────────────────────────────────────────────────────
import importlib.util, sys

def load_module(name, path):
    spec = importlib.util.spec_from_file_location(name, path)
    mod  = importlib.util.module_from_spec(spec)
    sys.modules[name] = mod
    spec.loader.exec_module(mod)
    return mod

BASE     = os.path.dirname(os.path.abspath(__file__))
PARSER   = load_module("parse_t12_xlsx",       os.path.join(BASE, "parse_t12_xlsx.py"))
POP      = load_module("populate_budget_model", os.path.join(BASE, "populate_budget_model.py"))
TEMPLATE = os.path.join(BASE, "Budget_Model.xlsx")

# ── Rent roll parser ──────────────────────────────────────────────────────────
def parse_date_val(v):
    if v is None: return None
    if hasattr(v, "date"): return v.date()
    if isinstance(v, str):
        for fmt in ("%m/%d/%Y", "%Y-%m-%d"):
            try: return datetime.strptime(v.strip(), fmt).date()
            except: pass
    return None

def parse_rent_roll(rr_bytes, total_units):
    wb = openpyxl.load_workbook(io.BytesIO(rr_bytes), data_only=True)
    ws = wb.worksheets[0]
    OCCUPIED = {"Occupied", "Occupied-NTV", "Occupied-NTVL"}
    VACANT   = {"Vacant", "Admin/Down"}

    # Yardi positional layout defaults
    unit_col, fp_col, sqft_col = 2, 3, 5
    status_col, mkt_col, act_col, lease_col = 6, 12, 18, 11
    header_row = 6

    unit_dict = {}
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if not row or len(row) < max(unit_col, fp_col) + 1: continue
        unit   = row[unit_col]
        fp     = row[fp_col]
        status = str(row[status_col] if status_col < len(row) else "").strip()
        if status in ("Pending renewal", "Applicant", "Pending resident"): continue
        if not unit or not fp: continue
        if str(fp).strip().lower() in ("floorplan", "unit type", "unit"): continue
        if unit in unit_dict: continue
        unit_dict[unit] = {
            "unit":        unit,
            "floorplan":   str(fp).strip(),
            "sqft":        int(row[sqft_col]) if sqft_col < len(row) and row[sqft_col] else 0,
            "status":      status,
            "market_rent": float(row[mkt_col]) if mkt_col < len(row) and row[mkt_col] else 0,
            "actual_rent": float(row[act_col]) if act_col < len(row) and row[act_col] else 0,
            "lease_end":   parse_date_val(row[lease_col] if lease_col < len(row) else None),
        }

    units = list(unit_dict.values())
    if not units:
        raise ValueError("No unit rows found in rent roll. Please check the file.")

    # Infer RR date from most recent past lease end date
    today = date.today()
    past  = [u["lease_end"] for u in units if u["lease_end"] and u["lease_end"] <= today]
    rr_date = max(past) if past else today

    def get_beds(fp):
        fp = fp.upper()
        if fp.startswith("C"): return 3
        if fp.startswith("B"): return 2
        return 1

    fp_data = defaultdict(lambda: {"sqfts":[],"mkt_rents":[],"actual_rents":[],"statuses":[]})
    for u in units:
        fp = u["floorplan"]
        fp_data[fp]["sqfts"].append(u["sqft"])
        fp_data[fp]["mkt_rents"].append(u["market_rent"])
        if u["actual_rent"] > 0: fp_data[fp]["actual_rents"].append(u["actual_rent"])
        fp_data[fp]["statuses"].append(u["status"])

    fp_rows = []
    for fp in sorted(fp_data.keys()):
        d    = fp_data[fp]
        n    = len(d["sqfts"])
        sqft = int(round(statistics.mean(d["sqfts"]))) if d["sqfts"] else 0
        mkt  = round(statistics.mean(d["mkt_rents"]), 2) if d["mkt_rents"] else 0
        ppsf = round(mkt / sqft, 4) if sqft else 0
        act  = round(statistics.mean(d["actual_rents"]), 2) if d["actual_rents"] else 0
        occ  = sum(1 for s in d["statuses"] if s in OCCUPIED)
        fp_rows.append({"unit_type":fp,"bedrooms":get_beds(fp),"unit_count":n,
                        "net_sf":sqft,"ppsf":ppsf,"market_rent":mkt,
                        "occupied_units":occ,"actual_rent":act})

    # Lease expirations
    window_start = date(rr_date.year + (rr_date.month // 12), (rr_date.month % 12) + 1, 1)
    months = []
    d2 = window_start
    for _ in range(12):
        last_day = calendar.monthrange(d2.year, d2.month)[1]
        months.append(date(d2.year, d2.month, last_day))
        d2 = date(d2.year + (d2.month // 12), (d2.month % 12) + 1, 1)

    monthly_exp = {m: 0 for m in months}
    excluded = beyond = 0
    for u in units:
        if u["status"] not in OCCUPIED: continue
        le = u["lease_end"]
        if not le: continue
        if le < window_start: excluded += 1; continue
        if le > months[-1]:   beyond += 1;   continue
        for m in months:
            if le <= m: monthly_exp[m] += 1; break

    total_occ = sum(1 for u in units if u["status"] in OCCUPIED)
    total_vac = sum(1 for u in units if u["status"] in VACANT)
    wavg_mkt  = sum(r["market_rent"]*r["unit_count"] for r in fp_rows) / len(units) if units else 0

    summary = {
        "parsed_units": len(units), "floorplans": len(fp_rows),
        "occupied": total_occ, "vacant": total_vac,
        "occ_pct": total_occ/len(units) if units else 0,
        "wavg_mkt_rent": wavg_mkt,
        "expirations_12mo": sum(monthly_exp.values()),
        "excluded": excluded, "beyond": beyond,
        "rr_date": rr_date, "window_start": window_start, "months": months,
    }

    warns = []
    if total_units and abs(len(units) - total_units) > 5:
        warns.append(f"Rent roll parsed {len(units)} units but you entered {total_units}. "
                     "Check for duplicate or missing rows.")
    return fp_rows, monthly_exp, window_start, months, summary, warns

# ── Main pipeline ─────────────────────────────────────────────────────────────
def run_pipeline(t12_bytes, rr_bytes, property_name, total_units, t12_sheet):
    warnings, validation = [], {}

    # Parse T-12
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        f.write(t12_bytes); t12_tmp = f.name
    try:
        items = PARSER.parse_t12(t12_tmp, sheet_index=t12_sheet)
    finally:
        os.unlink(t12_tmp)

    rev = [x for x in items if x["section"] == "Revenue"]
    exp = [x for x in items if x["section"] == "Expenses"]
    validation["rev_rows"]   = len(rev)
    validation["exp_rows"]   = len(exp)
    validation["rev_total"]  = sum(x["total"] for x in rev)
    validation["exp_total"]  = sum(x["total"] for x in exp)

    # Expense section breakdown
    by_sec = defaultdict(float)
    for x in exp: by_sec[x["t12_section"]] += x["total"]
    validation["by_section"] = dict(sorted(by_sec.items(), key=lambda x: -x[1]))

    wb = openpyxl.load_workbook(TEMPLATE)
    POP.populate_t12(wb["T-12 Inputs"], items, property_name=property_name)

    # Rent roll
    fp_rows = None
    if rr_bytes:
        try:
            fp_rows, monthly_exp, window_start, months, rr_sum, rr_warns = \
                parse_rent_roll(rr_bytes, total_units)
            warnings.extend(rr_warns)

            ws_rs = wb["Rent Schedule"]
            for i, fp in enumerate(fp_rows):
                r = 4 + i
                ws_rs.cell(r, 3).value  = fp["unit_type"]
                ws_rs.cell(r, 4).value  = fp["bedrooms"]
                ws_rs.cell(r, 5).value  = fp["unit_count"]
                ws_rs.cell(r, 6).value  = fp["net_sf"]
                ws_rs.cell(r, 7).value  = fp["ppsf"];   ws_rs.cell(r,7).number_format = "#,##0.00"
                ws_rs.cell(r, 8).value  = fp["occupied_units"]
                ws_rs.cell(r, 18).value = fp["market_rent"]; ws_rs.cell(r,18).number_format = "$#,##0.00"
                ws_rs.cell(r, 21).value = fp["actual_rent"];  ws_rs.cell(r,21).number_format = "$#,##0.00"

            ws_abs = wb["Absorption Schedule"]
            ws_abs["B6"] = window_start
            for i, m in enumerate(months):
                ws_abs.cell(row=27, column=9 + i).value = monthly_exp[m]

            validation["rr_units"]   = rr_sum["parsed_units"]
            validation["rr_fps"]     = rr_sum["floorplans"]
            validation["rr_occ_pct"] = rr_sum["occ_pct"]
            validation["rr_wavg"]    = rr_sum["wavg_mkt_rent"]
            validation["rr_exp12"]   = rr_sum["expirations_12mo"]
            validation["rr_date"]    = rr_sum["rr_date"]

        except Exception as e:
            warnings.append(f"Rent roll error: {e}. Rent Schedule not populated.")

    POP.set_units(wb, total_units, skip_rent_schedule=(fp_rows is not None))

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        out_tmp = f.name
    POP.save_with_zip_surgery(wb, TEMPLATE, out_tmp)
    with open(out_tmp, "rb") as f:
        output_bytes = f.read()
    os.unlink(out_tmp)

    return output_bytes, validation, warnings

# ─────────────────────────────────────────────────────────────────────────────
# UI
# ─────────────────────────────────────────────────────────────────────────────

st.markdown("""
<div class="wordmark">
  <span style="font-size:20px">🏢</span>
  <span class="wordmark-name">Budget Model Importer</span>
  <span class="wordmark-tag">Multifamily</span>
</div>
""", unsafe_allow_html=True)

if not os.path.exists(TEMPLATE):
    st.error(f"**Budget_Model.xlsx not found** in `{BASE}`. Add the template file to proceed.")
    st.stop()

# Property details
st.markdown('<div class="step-label">Property details</div>', unsafe_allow_html=True)
col1, col2 = st.columns([3, 1])
with col1:
    property_name = st.text_input("Property name", placeholder="e.g. Market Station",
                                   label_visibility="collapsed")
with col2:
    total_units = st.number_input("Units", min_value=1, max_value=9999,
                                   value=None, placeholder="329",
                                   label_visibility="collapsed")

# T-12 upload
st.markdown('<div class="step-label">T-12 operating statement (.xlsx)</div>', unsafe_allow_html=True)
t12_file = st.file_uploader("T-12", type=["xlsx"], label_visibility="collapsed", key="t12")

t12_sheet = 0
if t12_file:
    try:
        wb_peek = openpyxl.load_workbook(io.BytesIO(t12_file.read()), read_only=True)
        sheets  = wb_peek.sheetnames; t12_file.seek(0)
        if len(sheets) > 1:
            sel = st.selectbox("Data sheet", sheets,
                               index=sheets.index("Report1") if "Report1" in sheets else 0)
            t12_sheet = sheets.index(sel)
    except: t12_file.seek(0)

# Rent roll upload
st.markdown('<div class="step-label">Rent roll <span style="color:#9CA3AF;font-weight:400;text-transform:none;letter-spacing:0">(optional)</span></div>',
            unsafe_allow_html=True)
rr_file = st.file_uploader("Rent roll", type=["xlsx"], label_visibility="collapsed", key="rr")

st.markdown("<br>", unsafe_allow_html=True)

ready = bool(t12_file and property_name and total_units)
if not ready:
    missing = [x for x, c in [("property name", property_name),
                                ("unit count", total_units),
                                ("T-12 file", t12_file)] if not c]
    if missing:
        st.caption(f"Waiting for: {', '.join(missing)}")

generate = st.button("Generate budget model →", disabled=not ready,
                     use_container_width=True, type="primary")

# ── Run ───────────────────────────────────────────────────────────────────────
if generate and ready:
    t12_bytes = t12_file.read()
    rr_bytes  = rr_file.read() if rr_file else None

    with st.spinner("Parsing and populating model…"):
        try:
            output_bytes, val, warnings = run_pipeline(
                t12_bytes, rr_bytes, property_name, int(total_units), t12_sheet)
            ok = True
        except Exception as e:
            st.error(f"**Error:** {e}")
            ok = False

    if ok:
        has_rr = "rr_units" in val
        note   = " · Rent Schedule populated" if has_rr else ""
        st.markdown(
            f'<div class="success-banner">✓ &nbsp; {property_name} — {int(total_units):,} units{note}</div>',
            unsafe_allow_html=True)

        safe = re.sub(r"[^\w\s-]", "", property_name).strip().replace(" ", "_")
        st.download_button(
            label=f"⬇  Download {safe}_Budget_Model.xlsx",
            data=output_bytes,
            file_name=f"{safe}_Budget_Model.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

        for w in warnings:
            st.markdown(f'<div class="warn-banner">⚠ &nbsp;{w}</div>', unsafe_allow_html=True)

        # Validation summary
        st.markdown("<br>", unsafe_allow_html=True)
        st.markdown('<div class="step-label">Validation</div>', unsafe_allow_html=True)

        def fm(v): return f"${v:,.0f}" if v else "—"
        def fp(v): return f"{v:.1%}" if v else "—"

        rows = [
            ("Total Revenue (T-12)",  fm(val.get("rev_total"))),
            ("Total Expenses (T-12)", fm(val.get("exp_total"))),
            ("Revenue line items",    str(val.get("rev_rows","—"))),
            ("Expense line items",    str(val.get("exp_rows","—"))),
        ]
        if has_rr:
            rows += [
                ("Rent roll units",       str(val.get("rr_units","—"))),
                ("Floorplans",            str(val.get("rr_fps","—"))),
                ("Occupancy",             fp(val.get("rr_occ_pct"))),
                ("Wtd avg market rent",   fm(val.get("rr_wavg"))),
                ("Lease expirations (12mo)", str(val.get("rr_exp12","—"))),
                ("Rent roll date",        str(val.get("rr_date","—"))),
            ]

        html = "".join(
            f'<div class="val-row"><span class="val-label">{l}</span>'
            f'<span class="val-value">{v}</span></div>'
            for l, v in rows)
        st.markdown(f'<div class="info-card">{html}</div>', unsafe_allow_html=True)

        # Expense section breakdown
        if val.get("by_section"):
            st.markdown('<div class="step-label">Expense sections (T-12)</div>',
                        unsafe_allow_html=True)
            sec_html = "".join(
                f'<div class="val-row"><span class="val-label">{s}</span>'
                f'<span class="val-value">{fm(a)}</span></div>'
                for s, a in val["by_section"].items())
            st.markdown(f'<div class="info-card">{sec_html}</div>', unsafe_allow_html=True)

st.markdown("""
<hr style="border:none;border-top:1px solid #E5E7EB;margin:3rem 0 1rem">
<p style="font-size:12px;color:#9CA3AF;text-align:center;margin:0">
  Budget Model Importer · Yardi / MRI format
</p>
""", unsafe_allow_html=True)
