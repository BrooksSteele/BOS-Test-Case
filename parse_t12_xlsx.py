"""
parse_t12_xlsx.py
-----------------
Reads a T-12 operating statement XLSX (Yardi / MRI style) and produces
the JSON array consumed by populate_budget_model.py.

Usage:
    python parse_t12_xlsx.py \
        --input  Sawyer_Yards_T12.xlsx \
        --output t12_parsed.json \
        [--sheet 0]               # sheet index or name, default 0

The parser:
  - Skips header/title rows (rows 1-5 in this format)
  - Identifies section headers (GL code ends in -000, no monthly data)
  - Skips subtotal rows (GL code ends in -099, -098, -199, -999)
  - Stops at NOI (GL 69999-099) — nothing below the line is included
  - Assigns a t12_section label to every data row based on the most
    recent section header
  - Maps each line item to a Proforma budget code using GL prefix rules
    (no cross-section reclassification — items stay in their T-12 section)
  - Strips leading spaces from description field

GL prefix -> t12_section -> budget_code mapping is defined in
SECTION_MAP and CODE_MAP at the top of the file. Extend these dicts
to support other chart-of-accounts structures.
"""

import argparse
import json
import os
import sys

import openpyxl


# ---------------------------------------------------------------------------
# Section header GL codes -> clean section name
# Used to assign t12_section to every data row that follows.
# ---------------------------------------------------------------------------
SECTION_NAME_MAP = {
    "40001-000": "Rental Income",
    "41030-000": "Rental Income",          # sub-section, same bucket
    "42000-000": "Corporate Housing Income",
    "43000-000": "Other Income - Residential",
    "50001-000": "Payroll & Benefits",
    "52000-000": "Repairs & Maintenance",  # parent of R&M sub-sections
    "52001-000": "Repairs & Maintenance",
    "52600-000": "Make-Ready / Redecorating",
    "52800-000": "Recreational Amenities",
    "53000-000": "Contract Services",
    "54000-000": "Advertising / Marketing",
    "58000-000": "General & Administrative",
    "58001-000": "General & Administrative",  # Office Expenses sub-section
    "58200-000": "General & Administrative",  # Other G&A sub-section
    "59000-000": "Utilities",
    "60000-000": "Management Fees",
    "62000-000": "Taxes",
    "63000-000": "Insurance",
}

# ---------------------------------------------------------------------------
# GL code -> Proforma budget_code
# Maps each individual data-row GL code to the correct Proforma dropdown code.
# Rules:
#   - Payroll items stay within payroll family
#   - R&M items stay as R&M (including recreational amenities)
#   - Contract Services items stay as Contract Services
#     (except Landscape -> Landscaping, Pest -> Pest, Trash -> R&M - Trash)
#   - Marketing items stay as Marketing / Locator Fees / Referral Fees /
#     Resident Retention
#   - G&A items stay within G&A family
#   - Utilities map to specific utility codes
#   - Revenue items map to Proforma revenue codes
# Any GL code not listed falls back to the section default (SECTION_CODE_DEFAULT).
# ---------------------------------------------------------------------------
GL_CODE_MAP = {
    # ── REVENUE ──────────────────────────────────────────────────────────────
    "41000-000": "GPR",
    "41010-000": "Loss to Lease",
    "41091-000": "Concessions",
    "41092-000": "Concessions",
    "41100-000": "Vacancy",
    "41110-000": "Model/Employee Discount",
    "41115-000": "Model/Employee Discount",
    "41120-000": "Model/Employee Discount",
    "41150-000": "Bad Debt",
    "41155-000": "Bad Debt",
    "42010-000": "Other Income",
    "43010-000": "Administrative Fees",
    "43016-600": "Amenity Fee",
    "43020-000": "Application Fees",
    "43055-000": "Cable Income",
    "43060-000": "Other Income",
    "43080-000": "Damages/Eviction Fees",
    "43097-000": "Damages/Eviction Fees",
    "43125-000": "Other Income",
    "43135-000": "Late Charges",
    "43145-000": "Termination/Transfer Fees",
    "43150-000": "Other Income",        # Legal Fees income = Other Income
    "43160-000": "Damages/Eviction Fees",
    "43170-000": "MTM Fees",
    "43180-000": "NSF Fees",
    "43185-000": "Package Fees",
    "43190-000": "Parking",
    "43200-000": "Pet",
    "43201-000": "Pet",
    "43213-000": "Other Income",
    "43215-000": "Other Income",        # Renter's Insurance — maps to Proforma 'Other Income'
    "43230-000": "MTM Fees",            # Short Term Premiums
    "43250-000": "Termination/Transfer Fees",
    "43258-000": "Utility Reimbursement",
    "43260-000": "Utility Reimbursement",
    "43261-000": "Pest",                # Pest Control Rebill income
    "43262-000": "Trash",
    "43263-000": "Trash",
    "43264-001": "Utility Reimbursement",
    "43267-000": "Other Income",
    "43290-000": "Other Income",

    # ── PAYROLL ───────────────────────────────────────────────────────────────
    "51010-000": "Salary - Property Management",
    "51020-000": "Salary - Property Management",
    "51024-000": "Payroll - Miscellaneous",  # Leasing Overtime
    "51030-000": "Leasing Bonuses",
    "51030-001": "Bonuses",
    "51040-000": "Salary - Maintenance ",    # trailing space matches Proforma
    "51040-001": "Payroll - Miscellaneous",  # Maintenance Overtime
    "51045-000": "Salary - Maintenance ",
    "51050-000": "Salary - Maintenance ",
    "51060-000": "Salary - Maintenance ",
    "51090-000": "Payroll - Burden",
    "51110-000": "Payroll - Burden",
    "51120-000": "Payroll - Burden",
    "51140-000": "Salary - Property Management",  # Centralization Fees
    "51160-000": "Salary - Maintenance ",

    # ── REPAIRS & MAINTENANCE (incl. Recreational Amenities) ─────────────────
    # All items in these T-12 sections map to R&M
    # (individual items can be recoded via col A dropdown after load)

    # ── CONTRACT SERVICES ────────────────────────────────────────────────────
    "53105-000": "Landscaping",    # Landscape Maintenance Contract
    "53140-000": "Pest",           # Pest Control Contract
    "53180-000": "R&M - Trash",    # Trash Removal Contract
    "53182-000": "R&M - Trash",    # Trash Removal - Door to Door Pickup

    # ── ADVERTISING / MARKETING ───────────────────────────────────────────────
    "54050-000": "Locator Fees",
    "54055-000": "Referral Fees",
    "54110-000": "Resident Retention",
    "54122-000": "Resident Retention",

    # ── G&A ───────────────────────────────────────────────────────────────────
    "58020-000": "Telephone",
    "58080-000": "Supplies",
    "58090-000": "Telephone",
    "58100-000": "Supplies",
    "58110-000": "Telephone",
    "58210-000": "Memberships",
    "58225-000": "Bank Charges",
    "58247-000": "G&A - Miscellaneous",
    "58250-000": "G&A - Miscellaneous",
    "58253-000": "G&A - Miscellaneous",
    "58260-000": "Eviction Charges",
    "58270-000": "Internet - GA",
    "58275-000": "Legal",
    "58290-000": "G&A - Miscellaneous",
    "58305-000": "Supplies",
    "58320-000": "G&A - Miscellaneous",

    # ── UTILITIES ─────────────────────────────────────────────────────────────
    "59020-000": "Electricity",
    "59040-000": "Electricity",
    "59070-000": "Gas",
    "59080-000": "Gas",
    "59100-000": "Utility Fees",
    "59110-000": "Water",

    # ── MANAGEMENT / TAXES / INSURANCE ───────────────────────────────────────
    "61030-000": "Management Fee",
    "62010-000": "Taxes",
    "63010-000": "Insurance",
}

# Default budget_code per t12_section for any GL code not in GL_CODE_MAP
SECTION_DEFAULT_CODE = {
    "Rental Income":                "GPR",
    "Corporate Housing Income":     "Other Income",
    "Other Income - Residential":   "Other Income",
    "Payroll & Benefits":           "Salary - Property Management",
    "Repairs & Maintenance":        "R&M",
    "Make-Ready / Redecorating":    "Make-Ready/Redec",
    "Recreational Amenities":       "R&M",
    "Contract Services":            "Contract Services",
    "Advertising / Marketing":      "Marketing",
    "General & Administrative":     "G&A",
    "Utilities":                    "Electricity",
    "Management Fees":              "Management Fee",
    "Taxes":                        "Taxes",
    "Insurance":                    "Insurance",
}

# GL codes that mark the end of operating data (NOI line)
NOI_GL_CODES = {"69999-099", "69999-090", "66999-199"}

# GL code suffix patterns that indicate subtotal/header rows to skip
SUBTOTAL_SUFFIXES = ("-099", "-098", "-199", "-999")


def is_subtotal(gl_code: str) -> bool:
    return any(gl_code.endswith(s) for s in SUBTOTAL_SUFFIXES)


def is_section_header(gl_code: str, has_monthly_data: bool) -> bool:
    return gl_code.endswith("-000") and not has_monthly_data


def clean_name(raw: str) -> str:
    """Strip leading spaces and normalise whitespace in description."""
    return " ".join(raw.split()) if raw else ""


def determine_section(gl_code: str, current_section: str) -> str:
    """Return the t12_section for a given GL code, or keep current."""
    return SECTION_NAME_MAP.get(gl_code, current_section)


def determine_section_type(current_section: str) -> str:
    """Return 'Revenue' or 'Expenses' based on section name."""
    revenue_sections = {
        "Rental Income",
        "Corporate Housing Income",
        "Other Income - Residential",
    }
    return "Revenue" if current_section in revenue_sections else "Expenses"


def get_budget_code(gl_code: str, t12_section: str) -> str:
    """Look up GL code, fall back to section default."""
    if gl_code in GL_CODE_MAP:
        return GL_CODE_MAP[gl_code]
    return SECTION_DEFAULT_CODE.get(t12_section, "Other Income")


def parse_t12(filepath: str, sheet_index=0) -> list:
    wb = openpyxl.load_workbook(filepath, data_only=True)

    if isinstance(sheet_index, int):
        ws = wb.worksheets[sheet_index]
    else:
        ws = wb[sheet_index]

    print(f"  Reading sheet: '{ws.title}' ({ws.max_row} rows)")

    items = []
    current_section = ""
    skipped_header = False

    for row_num, row in enumerate(ws.iter_rows(values_only=True), 1):
        gl_raw  = row[0]
        desc    = row[1]
        monthly = list(row[2:14])   # cols C-N = 12 months
        total   = row[14]           # col O = total

        # Skip blank rows
        if gl_raw is None and desc is None:
            continue

        gl_code = str(gl_raw).strip() if gl_raw else ""

        # Skip preamble rows (no GL code, just property name / period info)
        if not gl_code:
            continue

        # Stop at NOI
        if gl_code in NOI_GL_CODES:
            print(f"  Stopping at NOI row {row_num} ({gl_code})")
            break

        has_monthly = any(v is not None for v in monthly)

        # Section header row: update current section, don't emit a data row
        if is_section_header(gl_code, has_monthly):
            new_section = determine_section(gl_code, current_section)
            if new_section != current_section:
                current_section = new_section
            continue

        # Subtotal row: skip
        if is_subtotal(gl_code):
            continue

        # Data row: emit
        if has_monthly and desc:
            budget_code = get_budget_code(gl_code, current_section)
            section_type = determine_section_type(current_section)
            monthly_vals = [float(v) if v is not None else 0.0 for v in monthly]
            total_val = float(total) if total is not None else sum(monthly_vals)

            items.append({
                "section":       section_type,
                "t12_section":   current_section,
                "original_name": clean_name(str(desc)),
                "budget_code":   budget_code,
                "monthly":       monthly_vals,
                "total":         round(total_val, 2),
                "notes":         gl_code,
            })

    return items


def main():
    parser = argparse.ArgumentParser(
        description="Parse a T-12 XLSX into JSON for populate_budget_model.py"
    )
    parser.add_argument("--input",  required=True, help="Path to T-12 .xlsx file")
    parser.add_argument("--output", required=True, help="Path for output .json file")
    parser.add_argument("--sheet",  default=0,
                        help="Sheet index (int) or sheet name (default: 0)")
    args = parser.parse_args()

    if not os.path.exists(args.input):
        print(f"ERROR: {args.input} not found")
        sys.exit(1)

    # Allow --sheet as int or string
    try:
        sheet_ref = int(args.sheet)
    except ValueError:
        sheet_ref = args.sheet

    print(f"Parsing: {args.input}")
    items = parse_t12(args.input, sheet_index=sheet_ref)

    rev_items  = [x for x in items if x["section"] == "Revenue"]
    exp_items  = [x for x in items if x["section"] == "Expenses"]
    rev_total  = sum(x["total"] for x in rev_items)
    exp_total  = sum(x["total"] for x in exp_items)

    print(f"  Parsed {len(rev_items)} revenue rows  (total: ${rev_total:,.2f})")
    print(f"  Parsed {len(exp_items)} expense rows  (total: ${exp_total:,.2f})")

    # Section breakdown
    from collections import defaultdict
    by_section = defaultdict(float)
    for x in exp_items:
        by_section[x["t12_section"]] += x["total"]
    print("\n  Expense section totals:")
    for sec, total in by_section.items():
        print(f"    {sec:<40} ${total:>12,.2f}")

    with open(args.output, "w", encoding="utf-8") as f:
        json.dump(items, f, indent=2)

    print(f"\nWritten: {args.output} ({len(items)} items)")


if __name__ == "__main__":
    main()
