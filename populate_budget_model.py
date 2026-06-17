"""
populate_budget_model.py
------------------------
Populates the Budget_Model.xlsx T-12 Inputs sheet from JSON produced
by the AI upload tool.

The JSON must include a 't12_section' field on each item (the T-12's
own section header, e.g. "Payroll & Benefits", "Contract Services").
Items are written in original T-12 order with blank spacer rows
inserted between T-12 sections — no reordering or recoding.

Usage:
    python populate_budget_model.py \
        --model   Budget_Model.xlsx \
        --t12     t12_data.json \
        --units   325 \
        --output  Budget_Model_Populated.xlsx \
        --property "Bellrock Sawyer Yards"
"""

import argparse, copy, io, json, os, re, shutil, sys, tempfile, zipfile
from collections import OrderedDict

from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------
T12_SHEET         = "T-12 Inputs"
T12_PROPERTY_CELL = "C2"
T12_REV_START_ROW = 7
T12_REV_END_ROW   = 100
T12_EXP_START_ROW = 105
T12_EXP_END_ROW   = 430

T12_COL_CODE  = 1   # A — budget code (Proforma SUMPRODUCT key); dropdown here
T12_COL_NAME  = 2   # B — original T-12 line item name
T12_COL_NOTES = 3   # C — GL account / T-12 section label
T12_COL_M1    = 4   # D — first month
T12_COL_M12   = 15  # O — twelfth month
T12_COL_TOTAL = 17  # Q — total formula

T12_REV_NUM_FMT   = '_(* #,##0.00_);_(* \\(#,##0.00\\);_(* "-"??_);_(@_)'
T12_EXP_NUM_FMT   = '_("$"* #,##0.00_);_("$"* \\(#,##0.00\\);_("$"* "-"??_);_(@_)'
T12_TOTAL_NUM_FMT = '_("$"* #,##0.00_);_("$"* \\(#,##0.00\\);_("$"* "-"??_);_(@_)'

YELLOW_FILL = PatternFill("solid", fgColor="FFFFFFCC")

DROPDOWN_SHEET = "OpStatement List - Dropdowns"
REV_DROPDOWN   = f"='{DROPDOWN_SHEET}'!$I$5:$I$35"
EXP_DROPDOWN   = f"='{DROPDOWN_SHEET}'!$L$5:$L$53"

RENT_SHEET     = "Rent Schedule"
RENT_START_ROW = 4
RENT_MAX_ROW   = 30    # data rows 4-30; total units in E31 -> Proforma (Annual) C4
RENT_COL_TYPE  = 3
RENT_COL_BEDS  = 4
RENT_COL_COUNT = 5
RENT_COL_SF    = 6
RENT_COL_PPSF  = 7
RENT_COL_OCC   = 8
RENT_COL_WEEKS = 11

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def load_json(path):
    with open(path, encoding="utf-8") as f:
        return json.load(f)

def col_letter(n):
    result = ""
    while n:
        n, rem = divmod(n - 1, 26)
        result = chr(65 + rem) + result
    return result

def set_cell(ws, row, col, value, num_fmt=None, fill=None):
    cell = ws.cell(row=row, column=col)
    cell.value = value
    if num_fmt:
        cell.number_format = num_fmt
    if fill:
        cell.fill = copy.copy(fill)

def copy_row_format(ws, src_row, dst_row, col_start, col_end):
    for col in range(col_start, col_end + 1):
        src = ws.cell(row=src_row, column=col)
        dst = ws.cell(row=dst_row, column=col)
        if src.fill and src.fill.fill_type == "solid":
            dst.fill = copy.copy(src.fill)
        if src.font:
            dst.font = copy.copy(src.font)
        if src.number_format and src.number_format != "General":
            dst.number_format = src.number_format

def add_dropdown(ws, col_ltr, start_row, end_row, formula):
    dv = DataValidation(
        type="list", formula1=formula,
        allow_blank=True, showDropDown=False,
        showErrorMessage=False, showInputMessage=False,
    )
    dv.sqref = f"{col_ltr}{start_row}:{col_ltr}{end_row}"
    ws.add_data_validation(dv)

def write_spacer_row(ws, row):
    """Blank yellow row with a zero-sum formula so T-12 totals still work."""
    copy_row_format(ws, T12_EXP_START_ROW, row, 1, 17)
    for col in range(1, 18):
        ws.cell(row=row, column=col).value = None
    d, o = col_letter(T12_COL_M1), col_letter(T12_COL_M12)
    ws.cell(row=row, column=T12_COL_TOTAL).value = \
        f'=IFERROR(SUM({d}{row}:{o}{row}),"")'

def build_ordered_list(items):
    """
    Preserve original T-12 order but insert a None sentinel between
    distinct t12_section groups so spacer rows appear at section breaks.
    """
    result = []
    prev_section = None
    for item in items:
        sec = item.get("t12_section", "")
        if prev_section is not None and sec != prev_section:
            result.append(None)   # spacer sentinel
        result.append(item)
        prev_section = sec
    return result

# ---------------------------------------------------------------------------
# T-12 population
# ---------------------------------------------------------------------------
def populate_t12(ws, line_items, property_name=None):
    if property_name:
        ws[T12_PROPERTY_CELL] = property_name

    revenue  = [x for x in line_items if x.get("section", "").lower() == "revenue"]
    expenses = [x for x in line_items if x.get("section", "").lower() != "revenue"]

    # --- Revenue (no spacers needed — typically one continuous block) ---
    rev_row = T12_REV_START_ROW
    for item in revenue:
        if rev_row > T12_REV_END_ROW:
            print("  WARNING: revenue rows exceeded limit; truncating.")
            break
        copy_row_format(ws, T12_REV_START_ROW, rev_row, 1, 17)
        set_cell(ws, rev_row, T12_COL_CODE,  item.get("budget_code", ""),   fill=YELLOW_FILL)
        set_cell(ws, rev_row, T12_COL_NAME,  item.get("original_name", ""), fill=YELLOW_FILL)
        notes = item.get("notes", "")
        if notes:
            set_cell(ws, rev_row, T12_COL_NOTES, notes, fill=YELLOW_FILL)
        monthly = (item.get("monthly") or [])
        monthly = (monthly + [0] * 12)[:12]
        for m, val in enumerate(monthly):
            set_cell(ws, rev_row, T12_COL_M1 + m, float(val or 0),
                     num_fmt=T12_REV_NUM_FMT, fill=YELLOW_FILL)
        d, o = col_letter(T12_COL_M1), col_letter(T12_COL_M12)
        set_cell(ws, rev_row, T12_COL_TOTAL,
                 f'=IFERROR(SUM({d}{rev_row}:{o}{rev_row}),"")',
                 num_fmt=T12_TOTAL_NUM_FMT)
        rev_row += 1

    # --- Expenses (spacers between T-12 sections) ---
    ordered = build_ordered_list(expenses)
    exp_row = T12_EXP_START_ROW
    spacer_count = 0

    for item in ordered:
        if exp_row > T12_EXP_END_ROW:
            print("  WARNING: expense rows exceeded limit; truncating.")
            break
        if item is None:
            write_spacer_row(ws, exp_row)
            exp_row += 1
            spacer_count += 1
            continue
        copy_row_format(ws, T12_EXP_START_ROW, exp_row, 1, 17)
        set_cell(ws, exp_row, T12_COL_CODE,  item.get("budget_code", ""),   fill=YELLOW_FILL)
        set_cell(ws, exp_row, T12_COL_NAME,  item.get("original_name", ""), fill=YELLOW_FILL)
        # Store T-12 section in col C so it's visible alongside the line item
        section_label = item.get("t12_section", item.get("notes", ""))
        if section_label:
            set_cell(ws, exp_row, T12_COL_NOTES, section_label, fill=YELLOW_FILL)
        monthly = (item.get("monthly") or [])
        monthly = (monthly + [0] * 12)[:12]
        for m, val in enumerate(monthly):
            set_cell(ws, exp_row, T12_COL_M1 + m, float(val or 0),
                     num_fmt=T12_EXP_NUM_FMT, fill=YELLOW_FILL)
        d, o = col_letter(T12_COL_M1), col_letter(T12_COL_M12)
        set_cell(ws, exp_row, T12_COL_TOTAL,
                 f'=IFERROR(SUM({d}{exp_row}:{o}{exp_row}),"")',
                 num_fmt=T12_TOTAL_NUM_FMT)
        exp_row += 1

    # Dropdowns on col A
    add_dropdown(ws, "A", T12_REV_START_ROW, T12_REV_END_ROW, REV_DROPDOWN)
    add_dropdown(ws, "A", T12_EXP_START_ROW, T12_EXP_END_ROW, EXP_DROPDOWN)

    exp_items = sum(1 for x in ordered if x is not None)
    print(f"  T-12: wrote {len(revenue)} revenue rows, "
          f"{exp_items} expense rows, {spacer_count} section spacers")
    print("  T-12 write complete.")

# ---------------------------------------------------------------------------
# Rent Schedule
# ---------------------------------------------------------------------------
def populate_rent_schedule(ws, unit_types):
    if not unit_types:
        return
    max_rows = RENT_MAX_ROW - RENT_START_ROW + 1
    if len(unit_types) > max_rows:
        unit_types = unit_types[:max_rows]
    print(f"  Rent Schedule: writing {len(unit_types)} unit type rows")
    for i, ut in enumerate(unit_types):
        row = RENT_START_ROW + i
        ws.cell(row=row, column=RENT_COL_TYPE).value  = ut.get("unit_type", "")
        if ut.get("bedrooms") is not None:
            ws.cell(row=row, column=RENT_COL_BEDS).value = int(ut["bedrooms"])
        if ut.get("unit_count") is not None:
            ws.cell(row=row, column=RENT_COL_COUNT).value = int(ut["unit_count"])
        sf = ut.get("net_sf")
        if sf is not None:
            ws.cell(row=row, column=RENT_COL_SF).value = float(sf)
        ppsf = ut.get("ppsf")
        if ppsf is None and ut.get("market_rent") and sf and float(sf) > 0:
            ppsf = float(ut["market_rent"]) / float(sf)
        if ppsf is not None:
            ws.cell(row=row, column=RENT_COL_PPSF).value = round(float(ppsf), 4)
            ws.cell(row=row, column=RENT_COL_PPSF).number_format = "#,##0.00"
        if ut.get("occupied_units") is not None:
            ws.cell(row=row, column=RENT_COL_OCC).value = int(ut["occupied_units"])
        weeks = ut.get("concessions_weeks", 0)
        ws.cell(row=row, column=RENT_COL_WEEKS).value = int(weeks) if weeks else 0
    print("  Rent Schedule write complete.")

def set_units(wb, units, skip_rent_schedule=False):
    # Only write to Rent Schedule E4 when no floorplan data is being written.
    # When floorplan rows are present, E31=SUM(E4:E30) totals correctly on its own.
    if not skip_rent_schedule and RENT_SHEET in wb.sheetnames:
        wb[RENT_SHEET].cell(row=RENT_START_ROW, column=RENT_COL_COUNT).value = int(units)
        print(f"  Units ({units}) → Rent Schedule E{RENT_START_ROW}")
    for sn in ["Payroll Schedule", "Payroll Schedule - Lease-Up"]:
        if sn in wb.sheetnames:
            wb[sn]["C2"] = int(units)
            print(f"  Units ({units}) → '{sn}'!C2")
    # B2 is the label "Total Number of Units"; the actual value cell is B3
    for sn in ["G&A Assumptions", "R&M Assumptions", "Utilities Assumptions",
               "Other Income Assumptions", "Marketing Stack", "Marketing Stack - Lease-Up"]:
        if sn in wb.sheetnames:
            wb[sn]["B3"] = int(units)
            print(f"  Units ({units}) → '{sn}'!B3")

# ---------------------------------------------------------------------------
# Zip-surgery save
# ---------------------------------------------------------------------------
def _build_sheet_map(template_path: str) -> dict:
    """
    Returns {sheet_name: 'xl/worksheets/sheetN.xml'} by reading
    workbook.xml and its .rels file directly from the zip.
    """
    with zipfile.ZipFile(template_path) as z:
        wb_xml   = z.read("xl/workbook.xml").decode("utf-8")
        rels_xml = z.read("xl/_rels/workbook.xml.rels").decode("utf-8")

    sheets = re.findall(
        r'<sheet name="([^"]+)" sheetId="\d+" r:id="(rId\d+)"', wb_xml
    )
    rels = dict(re.findall(
        r'Id="(rId\d+)"[^>]*Target="(worksheets/[^"]+)"', rels_xml
    ))
    result = {}
    for raw_name, rid in sheets:
        name = (raw_name.replace("&amp;", "&")
                        .replace("&lt;",  "<")
                        .replace("&gt;",  ">"))
        if rid in rels:
            result[name] = f"xl/{rels[rid]}"
    return result


def _extract_x14_block(xml_bytes: bytes) -> bytes:
    """
    Return the raw <extLst>...</extLst> block containing x14 dataValidations,
    or b'' if absent. The extLst wrapper carries the required namespace decls.
    """
    text = xml_bytes.decode("utf-8", errors="replace")
    # The x14 DVs live inside <extLst> near the end of the worksheet
    start = text.find("<extLst>")
    if start == -1:
        # Some sheets use the bare x14:dataValidations without extLst
        start = text.find("<x14:dataValidations")
        if start == -1:
            return b""
        end = text.find("</x14:dataValidations>", start)
        if end == -1:
            return b""
        end += len("</x14:dataValidations>")
        return text[start:end].encode("utf-8")

    end = text.find("</extLst>", start)
    if end == -1:
        return b""
    end += len("</extLst>")
    return text[start:end].encode("utf-8")


def _inject_x14_block(openpyxl_xml: bytes, x14_block: bytes,
                      template_xml: bytes) -> bytes:
    """
    Insert the extLst/x14 block into the openpyxl-saved XML just before
    </worksheet>. Also swaps the openpyxl <worksheet> opening tag for the
    template's (which carries xr, x14ac and other needed namespace decls).
    """
    if not x14_block:
        return openpyxl_xml

    text = openpyxl_xml.decode("utf-8")
    tpl  = template_xml.decode("utf-8", errors="replace")

    # Replace the openpyxl <worksheet ...> tag with the template's richer version
    tpl_ws_start = tpl.find("<worksheet ")
    tpl_ws_end   = tpl.find(">", tpl_ws_start) + 1
    tpl_ws_tag   = tpl[tpl_ws_start:tpl_ws_end]

    opx_ws_start = text.find("<worksheet ")
    opx_ws_end   = text.find(">", opx_ws_start) + 1
    text = text[:opx_ws_start] + tpl_ws_tag + text[opx_ws_end:]

    # Remove any existing extLst in the openpyxl output (shouldn't be there,
    # but clean up just in case)
    if "<extLst>" in text:
        es = text.find("<extLst>")
        ee = text.find("</extLst>", es) + len("</extLst>")
        text = text[:es] + text[ee:]

    # Inject just before </worksheet>
    close_tag = "</worksheet>"
    if close_tag in text:
        text = text.replace(close_tag, x14_block.decode("utf-8") + close_tag, 1)

    return text.encode("utf-8")


def save_with_zip_surgery(wb, template_path: str, output_path: str):
    """
    Save the workbook while preserving all x14 extended data validations.

    Strategy:
      1. Save the openpyxl workbook to a temp file (gives us correct
         sharedStrings, styles, and all formula strings)
      2. For each sheet we modified, extract the x14 extLst block from
         the TEMPLATE and inject it back into the openpyxl-saved sheet XML
      3. Write the final zip from the openpyxl output with those patches

    This keeps sharedStrings, styles, and all inter-file references consistent
    (no mixing of template and openpyxl package files).

    fullCalcOnLoad=1 is already set in the workbook, so Excel recalculates
    all formulas on open — cached values are not needed.
    """
    MODIFIED_SHEETS = {
        "T-12 Inputs",
        "Rent Schedule",
        "Payroll Schedule",
        "Payroll Schedule - Lease-Up",
        "G&A Assumptions",
        "R&M Assumptions",
        "Utilities Assumptions",
        "Other Income Assumptions",
        "Marketing Stack",
        "Marketing Stack - Lease-Up",
        # Proforma and projection sheets with x14 DVs
        "Proforma (Annual)",
        "5-Year Projection",
        "Operating Budget - Monthly",
        "Lease-Up Proforma",
    }

    sheet_map = _build_sheet_map(template_path)

    # Save openpyxl workbook to temp
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp_path = tmp.name
    wb.save(tmp_path)

    try:
        # Read x14 extLst blocks from the template for sheets we care about
        x14_blocks = {}
        with zipfile.ZipFile(template_path) as z_tpl:
            for sheet_name in MODIFIED_SHEETS:
                xml_path = sheet_map.get(sheet_name)
                if xml_path:
                    tpl_xml = z_tpl.read(xml_path)
                    block = _extract_x14_block(tpl_xml)
                    if block:
                        x14_blocks[xml_path] = (block, tpl_xml)

        # Build final zip: start from openpyxl output, patch modified sheets
        with zipfile.ZipFile(tmp_path) as z_src:
            src_files = {name: z_src.read(name) for name in z_src.namelist()}

        with zipfile.ZipFile(output_path, "w", zipfile.ZIP_DEFLATED) as z_out:
            for name, data in src_files.items():
                if name in x14_blocks:
                    block, tpl_xml = x14_blocks[name]
                    data = _inject_x14_block(data, block, tpl_xml)
                z_out.writestr(name, data)

    finally:
        os.unlink(tmp_path)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--model",    required=True)
    parser.add_argument("--t12",      default=None)
    parser.add_argument("--rent",     default=None)
    parser.add_argument("--units",    default=None, type=int)
    parser.add_argument("--output",   required=True)
    parser.add_argument("--property", default=None, dest="property_name")
    args = parser.parse_args()

    if not os.path.exists(args.model):
        print(f"ERROR: {args.model} not found"); sys.exit(1)
    if not args.t12 and not args.rent:
        print("ERROR: provide --t12 and/or --rent"); sys.exit(1)

    print(f"Loading model: {args.model}")
    wb = load_workbook(args.model)

    if args.t12:
        if not os.path.exists(args.t12):
            print(f"ERROR: {args.t12} not found"); sys.exit(1)
        data = load_json(args.t12)
        items = data if isinstance(data, list) else data.get("line_items", [])
        populate_t12(wb[T12_SHEET], items, property_name=args.property_name)

    if args.rent:
        if not os.path.exists(args.rent):
            print(f"ERROR: {args.rent} not found"); sys.exit(1)
        data = load_json(args.rent)
        unit_types = data if isinstance(data, list) else data.get("unit_types", [])
        populate_rent_schedule(wb[RENT_SHEET], unit_types)

    if args.units:
        set_units(wb, args.units)

    print(f"Saving: {args.output}")
    save_with_zip_surgery(wb, args.model, args.output)
    print("Done.")

if __name__ == "__main__":
    main()
